import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from win32com.client import Dispatch
import os


# Step 1: Sample Data
data = {
    "Asset Name": ["Asset 1", "Asset 2", "Asset 3"],
    "Month 1": [15, 30, 40],
    "Month 2": [5, 35, 25],
}
df = pd.DataFrame(data)
df["Total"] = df["Month 1"] + df["Month 2"]

# Step 2: Create Workbook and Sheets
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"
ws_details = wb.create_sheet("Details")

# Step 3: Write DataFrame to 'Summary' sheet
for row in dataframe_to_rows(df, index=False, header=True):
    ws_summary.append(row)

# Step 4: Bold Header & Autofit Columns
for cell in ws_summary[1]:
    cell.font = Font(bold=True)

for col in ws_summary.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_summary.column_dimensions[col[0].column_letter].width = max_length + 2

# Step 5: Add Excel Table
table = Table(displayName="AssetTable", ref=f"A1:D{len(df)+1}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws_summary.add_table(table)

# Step 6: Add Column Chart for "Total"
chart = BarChart()
chart.title = "Total Assets Overview"
chart.y_axis.title = 'Total Value'
chart.x_axis.title = 'Asset Name'

data_ref = Reference(ws_summary, min_col=4, min_row=1, max_row=len(df)+1)  # "Total" column
cats_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=len(df)+1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4
ws_summary.add_chart(chart, "F2")

# Step 7: Add Conditional Formatting to highlight high Total (>60)
ws_summary.conditional_formatting.add(
    f"D2:D{len(df)+1}",
    CellIsRule(operator='greaterThan', formula=['60'], stopIfTrue=True,
               fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
)

# Step 8: Add Details Sheet with Calculations
ws_details.append(["Metric", "Value"])
ws_details.append(["Total Assets", f"=COUNTA(Summary!A2:A{len(df)+1})"])
ws_details.append(["Combined Total", f"=SUM(Summary!D2:D{len(df)+1})"])
for col in ws_details.columns:
    ws_details.column_dimensions[col[0].column_letter].width = 20
for cell in ws_details[1]:
    cell.font = Font(bold=True)

# Step 9: Save file
file_path = os.path.abspath("excelauto_output_advanced.xlsx")
wb.save(file_path)
print(f"✅ Advanced Excel file saved at: {file_path}")

# Step 10: Open in Excel
excel = Dispatch("Excel.Application")
excel.Visible = True
excel.Workbooks.Open(file_path)

# Optional: If you want to keep the script running until Excel is closed
# input("Press Enter to exit...")       # Uncomment this line if you want to pause the script


# Ensure the Excel application is visible and the workbook is opened
# Note: The above code requires the 'openpyxl' and 'pandas' libraries to be installed.  
